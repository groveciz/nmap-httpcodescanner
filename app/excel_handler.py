"""
Excel File Handler - Read uploads and write results
"""
import openpyxl
from openpyxl import Workbook
from typing import List, Dict
import os
from app.config import RESULTS_DIR


def extract_subdomain_zone(domain: str) -> tuple:
    """
    Extract subdomain and zone from a full domain name.
    
    Examples:
        api.example.com -> ('api', 'example.com')
        www.api.example.com -> ('www.api', 'example.com')
        example.com -> ('@', 'example.com')
    """
    parts = domain.split('.')
    
    if len(parts) <= 2:
        # example.com or example.co.uk style
        return '@', domain
    else:
        # Assume last 2 parts are the zone (works for .com, .net, .ru, etc.)
        # For .co.uk style domains, this won't be perfect but covers most cases
        zone = '.'.join(parts[-2:])
        subdomain = '.'.join(parts[:-2])
        return subdomain if subdomain else '@', zone


def read_excel(file_path: str) -> List[Dict]:
    """
    Read Excel file and extract domain/IP data.
    Auto-detects format:
    
    OLD FORMAT (2 columns):
    - Column A: Domain
    - Column B: IP Address
    
    NEW FORMAT (4 columns):
    - Column A: Subdomain
    - Column B: Zone
    - Column C: Domain
    - Column D: IP Address
    
    Returns:
        List of dicts with 'subdomain', 'zone', 'domain', and 'ip' keys
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    items = []
    
    # Detect format by checking first row column count
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    
    if not first_row:
        wb.close()
        return items
    
    # Count non-empty columns in first row
    col_count = sum(1 for cell in first_row if cell is not None)
    
    # Determine format: 4+ columns = new format, else old format
    is_new_format = col_count >= 4
    
    for row in ws.iter_rows(min_row=1, values_only=True):
        if is_new_format:
            # NEW FORMAT: Subdomain, Zone, Domain, IP
            subdomain = row[0] if len(row) > 0 and row[0] else "@"
            zone = row[1] if len(row) > 1 else None
            domain = row[2] if len(row) > 2 else None
            ip = row[3] if len(row) > 3 else None
        else:
            # OLD FORMAT: Domain, IP
            domain = row[0] if len(row) > 0 else None
            ip = row[1] if len(row) > 1 else None
            
            # Extract subdomain and zone from domain
            if domain:
                subdomain, zone = extract_subdomain_zone(str(domain).strip())
            else:
                subdomain, zone = None, None
        
        if domain and ip:
            items.append({
                "subdomain": str(subdomain).strip() if subdomain else "@",
                "zone": str(zone).strip() if zone else "",
                "domain": str(domain).strip(),
                "ip": str(ip).strip()
            })
    
    wb.close()
    return items


def write_excel(results: List[Dict], output_path: str) -> str:
    """
    Write scan results to Excel file.
    
    Output format:
    - Column A: Domain
    - Column B: IP
    - Column C: Ports
    - Column D: HTTP Status
    - Column E: HTTPS Status
    
    Returns:
        Path to created file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Scan Results"
    
    # Headers
    headers = ["Domain", "IP", "Ports", "HTTP Status", "HTTPS Status"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Data rows
    for row_idx, item in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=item.get("domain", ""))
        ws.cell(row=row_idx, column=2, value=item.get("ip", ""))
        ws.cell(row=row_idx, column=3, value=item.get("ports", ""))
        ws.cell(row=row_idx, column=4, value=item.get("http_status", ""))
        ws.cell(row=row_idx, column=5, value=item.get("https_status", ""))
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_path)
    wb.close()
    
    return output_path


def get_unique_ips(items: List[Dict]) -> List[str]:
    """
    Extract unique IP addresses from items list.
    """
    seen = set()
    unique = []
    for item in items:
        ip = item.get("ip")
        if ip and ip not in seen:
            seen.add(ip)
            unique.append(ip)
    return unique


def write_cloudflare_excel(results: List[Dict], output_path: str) -> str:
    """
    Write Cloudflare scan results to Excel file with zone info.
    
    Output format:
    - Column A: Subdomain
    - Column B: Zone
    - Column C: Domain
    - Column D: IP
    - Column E: Ports
    - Column F: HTTP Status
    - Column G: HTTPS Status
    
    Returns:
        Path to created file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Cloudflare Scan Results"
    
    # Headers
    headers = ["Subdomain", "Zone", "Domain", "IP", "Ports", "HTTP Status", "HTTPS Status"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Data rows
    for row_idx, item in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=item.get("subdomain", ""))
        ws.cell(row=row_idx, column=2, value=item.get("zone", ""))
        ws.cell(row=row_idx, column=3, value=item.get("domain", ""))
        ws.cell(row=row_idx, column=4, value=item.get("ip", ""))
        ws.cell(row=row_idx, column=5, value=item.get("ports", ""))
        ws.cell(row=row_idx, column=6, value=item.get("http_status", ""))
        ws.cell(row=row_idx, column=7, value=item.get("https_status", ""))
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_path)
    wb.close()
    
    return output_path

